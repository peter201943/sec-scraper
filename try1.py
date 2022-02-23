"""
2022 MIT License Drexel University
"""

import  requests
import  json
import  re
from    openpyxl        import load_workbook
from    openpyxl.styles import Alignment
from    bs4             import BeautifulSoup
from    ratelimit       import limits, RateLimitException, sleep_and_retry

WORKBOOK_NAME           = "final.xlsx"
WORKSHEET_NAME          = "export"
COLUMN_MAIN             = 1
COLUMN_SEC_LINK         = 9
COLUMN_10K_LINK         = 11
COLUMN_D_WORDCOUNT      = 12
COLUMN_D_SENTENCES      = 13
ROW_START               = 2
TEN_SECONDS             = 10 # Out of decency, using 10 seconds as opposed to 1 second long wait
MAX_CALLS_PER_SECOND    = 10
CHARACTER_SEARCH_RANGE  = 100
REGEX                   = 'divers' # for now, just using a simple search string
HEADERS                 = json.load(open("secrets.json"))["sec_request_headers"]

class SecLink():
  """
  Bad: `https://www.sec.gov/ix?doc=/Archives/edgar/data/1555280/000155528021000098/zts-20201231.htm`
  Fix: `https://www.sec.gov/Archives/edgar/data/1555280/000155528021000098/zts-20201231.htm`
  """
  def __init__(self,address:str = None):
    self.fixed = False
    if address is not None:
      self.address = address
      self.fix()
    else:
      self.address = ""
  def fix(self):
    if "ix?doc=/" in self.address:
      self.address = "https://www.sec.gov/" + self.address.split("ix?doc=/")[1]
    self.fixed = True
  def __str__(self):
    return self.address
  def __repr__(self):
    return f"<SecLink ({'fixed' if self.fixed else 'broken'}) \"{self.address}\">"

def test_read(wb=WORKBOOK_NAME, ws=WORKSHEET_NAME, idc=COLUMN_MAIN, target=COLUMN_SEC_LINK):
  workbook = load_workbook(wb)
  worksheet = workbook[ws]
  links = []
  for row_id in range(0,len(worksheet[idc])):
    next_row = row_id + ROW_START
    next_link = worksheet.cell(column=target,row=next_row).value
    if isinstance(next_link, str) and len(next_link) > 5:
      links.append(next_link)
  return(links)

def test_write(wb=WORKBOOK_NAME, ws=WORKSHEET_NAME):
  workbook = load_workbook(wb)
  worksheet = workbook[ws]
  # for row in safe_read(stats_path): # FIXME replace
  #   try:
  #     worksheet.cell(
  #       column  = 11, 
  #       row     = int(row['row_id']), 
  #       value   = "i can haz cheezburgr\ncat_salad_outrage.jpeg"
  #     ).alignment = Alignment(wrapText=True)
  #   except Exception as e:
  #     workbook.save(final_path)
  #     raise e
  workbook.save(wb)

def test_requests(headers=HEADERS, link="https://www.sec.gov/Archives/edgar/data/1555280/000155528021000098/zts-20201231.htm"):
  resp = requests.get(link, headers=headers)
  html = resp.text
  soup = BeautifulSoup(html, "html.parser")
  print(soup.body.get_text().strip())

def test_link():
  test_link = SecLink("https://www.sec.gov/ix?doc=/Archives/edgar/data/1555280/000155528021000098/zts-20201231.htm")
  print(f"test_link: {repr(test_link)}")

@sleep_and_retry
@limits(calls=MAX_CALLS_PER_SECOND, period=TEN_SECONDS)
def get_page_rate_limited(link:str, headers=HEADERS) -> BeautifulSoup:
  resp = requests.get(link, headers=headers)
  html = resp.text
  return BeautifulSoup(html, "html.parser")

def get_dir_10k_link(page_dir:BeautifulSoup) -> str:
  link_10k = ""
  for table_row in page_dir.find_all("tr"):
    try:
      if table_row.find_all("td")[1].string in ["10-K", "10K", "10k", "10-k"]:
        link_10k = SecLink(table_row.find_all('td')[2].a.get('href'))
    except:
      continue
  return link_10k

def get_diversity_instances(plaintext:str) -> list:
  min_distance = 0
  max_distance = len(plaintext)
  places = (match.start() for match in re.finditer(REGEX, plaintext))
  sentences = []
  for place in places:
    start = max(place - CHARACTER_SEARCH_RANGE, min_distance)
    stop  = min(place + CHARACTER_SEARCH_RANGE, max_distance)
    new_sentence = plaintext[start:stop]
    sentences.append(new_sentence)
  return sentences

def test_grabbing():
  sentences = get_diversity_instances(
    get_page_rate_limited(
      SecLink("https://www.sec.gov/ix?doc=/Archives/edgar/data/1555280/000155528021000098/zts-20201231.htm")
    ).body.get_text().strip()
  )
  print(f"instances: {len(sentences)}")
  print("sentences:")
  for sentence in sentences:
    print(f"- {sentence}")

def test_10k_link():
  with open("dir_0001555280-21-000098.htm") as local_html:
    link = get_dir_10k_link(BeautifulSoup(local_html, "html.parser"))
    print(link)

if __name__ == "__main__":
  # test_write()
  # test_read()
  # test_requests()
  # test_link()
  # test_requests()
  # access_rate_limited_api(SecLink("https://www.sec.gov/ix?doc=/Archives/edgar/data/1555280/000155528021000098/zts-20201231.htm"))
  # test_grabbing()
  test_10k_link()
  pass
