"""
2022 MIT License Drexel University
"""

import  requests
import  json
import  re
import  logging
from    openpyxl        import  load_workbook
from    openpyxl.styles import  Alignment
from    bs4             import  BeautifulSoup
from    ratelimit       import  limits, RateLimitException, sleep_and_retry
from    sec_scraper     import  *

def test_link():
  test_link = SecLink("https://www.sec.gov/ix?doc=/Archives/edgar/data/1555280/000155528021000098/zts-20201231.htm")
  print(f"test_link: {repr(test_link)}")

def test_requests(headers=HEADERS, link="https://www.sec.gov/Archives/edgar/data/1555280/000155528021000098/zts-20201231.htm"):
  resp = requests.get(link, headers=headers)
  html = resp.text
  soup = BeautifulSoup(html, "html.parser")
  print(soup.body.get_text().strip())

def test_grabbing():
  sentences = get_diversity_instances(
    get_page_rate_limited(
      SecLink("https://www.sec.gov/ix?doc=/Archives/edgar/data/1555280/000155528021000098/zts-20201231.htm")
    ).body.get_text().strip()
  )
  print(f"instances: {len(sentences)}")
  print("sentences:")
  for sentence in sentences:
    print(f"- {sentence}")

def test_read(wb=WORKBOOK_NAME, ws=WORKSHEET_NAME, idc=COLUMN_MAIN, target=COLUMN_SEC_LINK):
  workbook = load_workbook(wb)
  worksheet = workbook[ws]
  links = []
  for row_id in range(0,len(worksheet[idc])):
    next_row = row_id + ROW_START
    next_link = worksheet.cell(column=target,row=next_row).value
    if isinstance(next_link, str) and len(next_link) > 5:
      links.append(next_link)
  return(links)

def test_write(wb=WORKBOOK_NAME, ws=WORKSHEET_NAME):
  workbook = load_workbook(wb)
  worksheet = workbook[ws]
  # for row in safe_read(stats_path): # FIXME replace
  #   try:
  #     worksheet.cell(
  #       column  = 11, 
  #       row     = int(row['row_id']), 
  #       value   = "i can haz cheezburgr\ncat_salad_outrage.jpeg"
  #     ).alignment = Alignment(wrapText=True)
  #   except Exception as e:
  #     workbook.save(final_path)
  #     raise e
  workbook.save(wb)

def test_10k_link():
  with open("dir_0001555280-21-000098.htm") as local_html:
    link = get_dir_10k_link(BeautifulSoup(local_html, "html.parser"))
    print(link)

def full_pipe_test():
  # using last item in excel file
  row_id = 4239
  # grab the sec-link (using function) (NEW) (finds link in sheet)
  dir_link = get_sheet_dir_link(row_id)
  # downloads dir page, loads page for processing, access sec.gov (using function)
  dir_page = get_page_rate_limited(dir_link)
  # get the doc_link (using function), extracts 10k link
  clean_10k_link = get_dir_10k_link(dir_page)
  # downloads 10k, download the doc (using function), Cleans up the 10k
  clean_10k = get_page_rate_limited(clean_10k_link).body.get_text().strip()
  # scan the doc (using function)
  sentences = get_diversity_instances(clean_10k)
  # write results (using function) (NEW)
  write_sentence_stats(4239,sentences)

if __name__ == "__main__":
  # test_link()
  # test_requests()
  # test_grabbing()
  # test_read()
  # test_write()
  # test_10k_link()
  # full_pipe_test()
  pass

