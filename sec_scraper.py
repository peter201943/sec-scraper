"""
MIT License

Copyright (c) 2022 Peter J. Mangelsdorf

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.
"""

## Imports
# Allows making HTTP connections over the Internet
import  requests
# Allows reading of JSON, a common plaintext storage medium
import  json
# "re" as in REGEX, or "Regular Expression", allows advanced searching of plaintext files
import  re
# Allows better error tracing, especially for long-running scripts
import  logging
# Allows reading and writing to Microsoft Excel Spreadsheets
from    openpyxl        import load_workbook
# Allows setting multiple lines in Excel cells
from    openpyxl.styles import Alignment
# Allows cleanup and navigation of HTML page files
from    bs4             import BeautifulSoup
# Prevents abuse of Internet services by limiting the number of calls made to them over time
from    ratelimit       import limits, RateLimitException, sleep_and_retry
# Allows for precise time operations
from    datetime        import datetime as __datetime
from    datetime        import timezone as __timezone
# Allows reading and writing of Filesystem folders across all operating systems
from    pathlib         import Path
# Allows introspection into errors at runtime
import  sys
import  traceback

## Logging

# On fresh install, make sure that the "logs" folder exists
Path("logs/").mkdir(parents=True, exist_ok=True)

def timename() -> str:
  """
  A simple function to generate Windows-compliant ISO datetime (ISO 8601) filenames
  Uses "Zulu" (Greenwhich Mean Time, +/- 0 hours) as opposed to local time
  """
  return str(__datetime.now(__timezone.utc)).replace(" ","T").replace("-00:00","").replace("+00:00","").replace(":",".") + "Z"

def start_logging():
  """
  Use CSV (Tabular Plaintext) files so non-technical users can inspect the logs
  Reports the:
  - time of an event ("when did this happen?/what order did events happen?")
  - type of event ("how important is this?")
  - location (in code) of an event ("who is invoking this event?")
  - details of an event ("why is this thing crashing?")
  Uses the last-written last-commit to indicate log-version/project-version
  (git rev-parse HEAD)
  """
  log_filename = f'logs/{timename()}.csv'
  logging.basicConfig(
    filename  = log_filename,
    encoding  = 'utf-8',
    level     = logging.DEBUG,
    format    = '"%(asctime)s",%(levelname)s,"%(filename)s.%(funcName)s:%(lineno)s","%(message)s"'
  )
  with open(log_filename, "a") as logfile:
    logfile.write('time,event,location,details\n')
  logging.info("project version: https://github.com/peter201943/sec-scraper/commit/94144b1d44ab83edfcc2d5f6701df5d436145a12")

# Initialize the logging
start_logging()

## Constants
WORKBOOK_NAME           = "kai-file.xlsx"
WORKSHEET_NAME          = "export"
COLUMN_MAIN             = 'A' # NOTICE that this is only used down in "update_workbook" for ONE CASE! (this is due to the inconsistent API)
COLUMN_SEC_LINK         = 9
COLUMN_D_WORDCOUNT      = 11
COLUMN_D_SENTENCES      = 12
COLUMN_CONAME           = 5
ROW_START               = 2
WAIT_SECONDS            = 10
MAX_CALLS_PER_SECOND    = 10
CHARACTER_SEARCH_RANGE  = 100
REGEX                   = re.compile(r'\bdiversity\b | \bdiverse\b',flags=re.IGNORECASE | re.VERBOSE)
HEADERS                 = json.load(open("secrets.json"))["sec_request_headers"]

# Log the constants
logging.debug(f"Variables: {dict(((k, globals()[k]) for k in ('WORKBOOK_NAME', 'WORKSHEET_NAME', 'COLUMN_MAIN', 'COLUMN_SEC_LINK', 'COLUMN_D_WORDCOUNT', 'COLUMN_D_SENTENCES', 'COLUMN_CONAME', 'ROW_START', 'WAIT_SECONDS', 'MAX_CALLS_PER_SECOND', 'CHARACTER_SEARCH_RANGE', 'REGEX')))}")

# For logging of errors, introduce a custom wrapper
def log_exceptions(task:callable):
  """
  Log exceptions as they occur (minimal details only)
  Does not interrupt the exception
  :param task: some function that may crash
  """
  def action(*args,**kwargs):
    try:
      return task(*args,**kwargs)
    except Exception as e:
      error_frame = traceback.extract_tb(sys.exc_info()[-1], 2)[1]
      logging.error(f"{error_frame.name}:{error_frame.lineno}: {repr(e)}")
      raise
  return action

## Utilities

class SecLink():
  """
  "Repairs" US SEC internal Edgar links from either fragments or "iXBRL" links to plain "Archive" links
  Bad:  https://www.sec.gov/ix?doc=/Archives/edgar/data/1555280/000155528021000098/zts-20201231.htm
  Bad:  Archives/edgar/data/1555280/000155528021000098/zts-20201231.htm
  Fix:  https://www.sec.gov/Archives/edgar/data/1555280/000155528021000098/zts-20201231.htm
  """
  @log_exceptions
  def __init__(self,address:str = None) -> None:
    """
    :param address: the URL from an SEC webpage
    """
    self.fixed = False
    if address is not None:
      self.address = address
      self.fix()
    else:
      self.address = ""
  @log_exceptions
  def fix(self) -> None:
    """
    With a set `address`, will attempt to fix the address and set it's status to "fixed"
    """
    old_address = self.address
    # If it looks like an iXBRL address, grab the second-half of it and make a new old-style address
    if "ix?doc=/" in self.address:
      self.address = "https://www.sec.gov/" + self.address.split("ix?doc=/")[1]
    # If it looks like a fragment of an address, make a new old-style address
    elif self.address[0:14] == "Archives/edgar":
      self.address = "https://www.sec.gov/" + self.address
    # If it looks like a fragment of an address, make a new old-style address
    elif self.address[0:15] == "/Archives/edgar":
      self.address = "https://www.sec.gov" + self.address
    # Do nothing if it looks like a regular address
    elif "https://www.sec.gov/" in self.address:
      pass
    else:
      raise ValueError(f"Given strange address, could not determine status: {old_address}")
    self.fixed = True
    logging.info(f"fixed link from: {old_address} to: {self.address}")
  def __str__(self) -> str:
    return self.address
  def __repr__(self) -> str:
    return f"<SecLink ({'fixed' if self.fixed else 'broken'}) \"{self.address}\">"

@log_exceptions
@sleep_and_retry
@limits(calls=MAX_CALLS_PER_SECOND, period=WAIT_SECONDS)
def get_page_rate_limited(link:str, headers=HEADERS) -> BeautifulSoup:
  """
  Downloads an HTML web page and returns it as a navigable Python object
  Will only start downloading the page if previous calls have not used up the call-budget (number-of-calls over time-in-seconds)
  :param link: the URL of the page to be downloaded
  :param headers: client information to be sent to the server hosting the webpage
  :return: Navigable Python representation of HTML
  """
  logging.info(f"downloading: {link}")
  resp = requests.get(link, headers=headers)
  html = resp.text
  return BeautifulSoup(html, "html.parser")

@log_exceptions
def get_sheet_dir_link(worksheet, row_id:int, target=COLUMN_SEC_LINK) -> str:
  """
  Grabs the directory-link of a company from the input worksheet
  :param worksheet: Source sheet to grab URL from
  :param row_id: Which company (row) to look for
  :param target: Which column is being used for the url
  :return: the directory link as a string
  """
  next_link = worksheet.cell(column=target,row=row_id).value
  if isinstance(next_link, str) and len(next_link) > 5:
    logging.info(f"found: {next_link}")
    return next_link
  else:
    raise Exception(f"encountered an entry with missing or invalid `sec_link`")

@log_exceptions
def get_dir_10k_link(page_dir:BeautifulSoup) -> str:
  """
  Looks for a link on a loaded SEC company directory page to a 10K form
  Also does other checks (filetype) and handling (weird naming)
  :param page_dir: Python representation of an HTML webpage
  :return: URL location of HTML 10K form, if it exists
  """
  link_10k = ""
  for table_row in page_dir.find_all("tr"):
    try:
      if table_row.find_all("td")[3].string in ["10-K", "10K", "10k", "10-k", "10-K Form", "10-K form", "10K Form"]:
        link_10k = SecLink(table_row.find_all('td')[2].a.get('href'))
        # Ignore PDFs
        if str(link_10k)[-4:] == ".pdf":
          continue
        # Only accept HTMLs
        elif str(link_10k)[-4:] == ".htm" or str(link_10k)[-5:] == ".html":
          break
        # If the link has no identifiable extension, drop it and log it (anomaly)
        else:
          logging.debug(f"found an unusual link: {link_10k}")
          link_10k = ""
    except:
      # Ignore pages without tables, etcetera
      continue
  if link_10k == "":
    raise Exception("encountered an entry with no linked 10K form")
  return link_10k

@log_exceptions
def get_diversity_instances(plaintext:str, regex:re.Pattern=REGEX, search_range:int=CHARACTER_SEARCH_RANGE) -> list:
  """
  Find all sentences which contain an instance of some word
  :param plaintext: The stripped (no HTML elements) representation of a webpage ("plaint text")
  :param regex: The "Regular Expression" to search with
  :param search_range: How many characters to the left and right of any found words to include into a "sentence"
  :return: list of all "sentences" with the diversity regex
  """
  min_distance = 0
  max_distance = len(plaintext)
  places = (match.start() for match in re.finditer(regex, plaintext))
  sentences = []
  for place in places:
    # Make sure the "sentence" does not accidentally select outside the range of the text itself
    start = max(place - search_range, min_distance)
    stop  = min(place + search_range, max_distance)
    new_sentence = plaintext[start:stop]
    sentences.append(new_sentence)
  return sentences

@log_exceptions
def write_sentence_stats(worksheet, row_id:int, sentences:list) -> None:
  """
  Calculates, formats, and stores sentence statistics into the worksheet
  NOTE does NOT save to file/disc!
  :param worksheet: Where to store the sentence statistics
  :param row_id: Which company the statistics are for
  :param sentences: The found sentences to extract statistics from
  """
  worksheet.cell(
    column  = COLUMN_D_WORDCOUNT,
    row     = row_id,
    value   = len(sentences)
  )
  worksheet.cell(
    column  = COLUMN_D_SENTENCES,
    row     = row_id,
    value   = "\n".join(sentences)
  ).alignment = Alignment(wrapText=True)
  logging.info(f"Added sentence statistics for row: {row_id}")

@log_exceptions
def cleanup_page(link:str) -> str:
  """
  Downloads the plaintext of a webpage
  :link: url of somewebpage
  :return: Plaintext body of a webpage
  """
  return get_page_rate_limited(
    link
  ).body.get_text(    # Get the text of a webpage
  ).strip(            # Remove all HTML elements
  ).replace("\n"," ") # Remove any newlines as well

@log_exceptions
def is_complete(worksheet, row_id:int, c_wordcount:int=COLUMN_D_WORDCOUNT, c_sentence:int=COLUMN_D_SENTENCES) -> bool:
  """
  Checks if a specific row has valid and complete sentence statistics
  :param worksheet: Loaded Python Worksheet representation
  :param row_id: Which company to check
  :param c_wordcount: Which column the wordcount is in
  :param c_sentence: Which column the sentence is in
  :return: Whether the row has valid statistics or not
  """
  d_wordcount = worksheet.cell(column = c_wordcount, row = row_id).value
  d_sentences = worksheet.cell(column = c_sentence, row = row_id).value
  # When a number is present
  if isinstance(d_wordcount,int):
    # Edge Case: Exactly no sentences have previously been found
    if d_wordcount == 0 and isinstance(d_sentences,str) and len(d_sentences) == 0:
      logging.debug(f"row {row_id} appears to already be complete")
      return True
    # Most Cases: Some sentences have been found
    elif d_wordcount > 0 and isinstance(d_sentences,str) and len(d_sentences) > 50:
      logging.debug(f"row {row_id} appears to already be complete")
      return True
    # Edge Case: Disagreement between `diversity_sentences` and `diversity_wordcount`
    else:
      logging.debug(f"row {row_id} has errors in `diversity_sentences` ({c_sentence}), will overwrite")
      return False
  # Empty or otherwise, rewrite
  else:
    logging.debug(f"row {row_id} has errors in `diversity_wordcount` ({d_wordcount}), will overwrite")
    return False

@log_exceptions
def update_workbook(row_ids:list=None, wb:str=WORKBOOK_NAME, ws:str=WORKSHEET_NAME, idc:int=COLUMN_MAIN, target:int=COLUMN_SEC_LINK, coname:int=COLUMN_CONAME) -> None:
  """
  Updates an entire notebook with statistics
  :param row_ids: A list of which rows to update (blank if want to update all)
  :param wb: Filename of the workbook to open
  :param ws: String page name of sheet in workbook to open
  :param idc: Column with the "primary identifier"/"primary key" of all items
  :param target: Which column to check for links
  :param coname: Which column contains the company name
  """
  logging.info("Started")
  workbook = load_workbook(wb)
  worksheet = workbook[ws]
  items = len(worksheet[idc])
  if not row_ids:
    row_ids = range(ROW_START,items + 1)
  if isinstance(row_ids,int):
    row_ids = [row_ids]
  logging.debug(f"Running script for {len(row_ids)} items")
  for row_id in row_ids:
    # Skip any rows outside the bounds of the file
    if row_id < ROW_START or row_id > items:
      logging.debug(f"Unusual row: {row_id}")
      continue
    company_name = worksheet.cell(column=coname,row=row_id).value
    logging.debug(f"NEXT row: {row_id} (\"{company_name}\")")
    try:
      if is_complete(worksheet,row_id):
        continue
      dir_link        = get_sheet_dir_link(worksheet,row_id,target)
      dir_page        = get_page_rate_limited(dir_link)
      clean_10k_link  = get_dir_10k_link(dir_page)
      clean_10k       = cleanup_page(clean_10k_link)
      sentences       = get_diversity_instances(clean_10k)
    except Exception as e:
      logging.error(f"SKIPPED row: {row_id}")
      continue
    try:
      write_sentence_stats(worksheet, row_id, sentences)
    except Exception as e:
      logging.critical(f"CRASHED on final statistics writing for row: {row_id}")
      logging.critical(f"Cancelling future writes, PLEASE INSPECT FILE MANUALLY FOR ERRORS")
      logging.critical("Exiting early")
      exit()
    workbook.save(wb)
    logging.debug(f"Saved Workbook")
  logging.info("Finished")

## Main

# Standard Python Meta-Manipulation to only execute the following code if it is invoked from the command line
if __name__ == "__main__":
  update_workbook()
  pass
